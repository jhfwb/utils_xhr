import os
import re
import openpyxl
from openpyxl.styles import Font
from _xhr_tool._utils.CsvTool import CsvTool
from _xhr_tool._utils.excelTool.csvFileOptionTool import csvFileTool
class ExcelTool:
    def __init__(self):
        pass
    def _createNewWorkbook(self):
        return openpyxl.Workbook()
    def _createNewSheet(self,workbook,sheetName):
        return workbook.create_sheet(sheetName)
    def _changeCellToStr(self,cell):
        if cell.hyperlink:
            cell.value="<href='"+cell.hyperlink.target+"'>"+cell.value
        return cell.value
    def _changeStrToCell(self,cell):
        attrss=re.findall(r'<(.*)>',cell.value)
        if len(attrss)>0:
            arr=[]
            attrs=attrss[0].split(' ')
            for attr in attrs:
                if attr.strip()!='':
                    arr.append(attr)
            for ar in arr:
                try:
                    tu=ar.split('=')
                    if tu[1].startswith("'") and tu[1].endswith("'"):
                        pass
                    elif tu[1].startswith("\"") and tu[1].endswith("\""):
                        pass
                    else:
                        raise ValueError()
                except:
                    raise ValueError("语法错误"+ar+"：无法处理该语句，请确保属性中有等号")
                if tu[0]=='href':############################################################################此处扩展
                    cell.hyperlink=tu[1][1:len(tu[1])-1]
                    cell.value=re.sub(r'<.*>','',cell.value)
                    cell.font=Font(bold=False,italic=False,underline="single",color='0000FF')
                else:
                    raise ValueError('语法错误：不存在'+tu[0]+'这个属性,目前只允许以下属性存在：href')


        return cell.value
    def _changeStyle(self,cell,fontStyle={}):
        """
        改变文字的样式fontStype
        underline:single | none _添加下划线
        fontColor: 0000FF _蓝色
        size: 11
        bold: True _是否加粗
        name: 宋体 _字体样式
        italic: True   _是否斜体
        strike: False _是否添加删除线
        :param cell:
        :param style:
        :return:
        """
        #设置默认样式
        if "size" in fontStyle.keys():
            pass
        cell.font = Font(size=fontStyle['size']|100, bold=True, italic=True, underline="single", color='0000FF')
        return cell
    def optionExecl(self,path='',sheetName='',datas=[],mode="",styleRemain=True):
        if not path.endswith('.xlsx'):
            raise NameError('命名错误,该execal文件必选以xlsx结尾')
        if mode=='w':
            wb=self._createNewWorkbook()
            sheet=self._createNewSheet(wb,sheetName)
            #表头的写入
            if len(datas)==0:
                del wb['Sheet']  # 删除默认表单
                wb.save(path)
                wb.close()
                return
            firstArr=list(datas[0].keys())
            for i in range(0,len(firstArr)):
                sheet.cell(row=1,column=i+1,value=firstArr[i])
            #表体的写入
            for j in range(0,len(datas)):
                for i in range(0,len(firstArr)):
                    cell=sheet.cell(row=j + 2, column=i + 1, value=datas[j].get(firstArr[i]))
                    #对value进行处理
                    if type(cell.value)==str:
                        cell=self._changeStrToCell(cell)
            del wb['Sheet']#删除默认表单
            wb.save(path)
            wb.close()
        elif mode=='r':
            wb = openpyxl.load_workbook(path)
            if sheetName=="":
                sh = wb[wb.sheetnames[0]]
            else:
                sh = wb[sheetName]
            rows_data = list(sh.rows)
            headLine=[]#头
            datas=[]
            for hc in rows_data.pop(0):
                headLine.append(hc.value)
            for row in rows_data:
                data={}
                for i in range(0,len(row)):
                    if styleRemain:
                        data.setdefault(headLine[i], self._changeCellToStr(row[i]))
                    else:
                        data.setdefault(headLine[i],row[i].value)
                datas.append(data)
            wb.close()
            return datas
        elif mode=='a':
            if not os.path.exists(path):
                self.optionExecl(path=path, sheetName=sheetName, datas=datas, mode="w")
                return
            oldDatas=[]
            try:
                oldDatas=self.optionExecl(path=path, sheetName=sheetName, mode="r", styleRemain=True)
            except IndexError:
                self.optionExecl(path=path, sheetName=sheetName, datas=datas, mode="w")
                return
            newDatas = oldDatas + datas
            self.optionExecl(path=path, sheetName=sheetName, datas=newDatas, mode="w")
    def getHeader(self,path='',sheetName=''):
        wb = openpyxl.load_workbook(path)
        if sheetName == "":
            sh = wb[wb.get_sheet_names()[0]]
        else:
            sh = wb[sheetName]
        rows_data = list(sh.rows)
        headLine=[]
        for hc in rows_data.pop(0):
            headLine.append(hc.value)
        return headLine
    def changeExeclToCsvFile(self,path="",encoding='utf-8'):
        """
        #将csv文件转成excel文件。目前只能转换第一个表格
        :param path:
        :param encoding:
        :return:
        """
        datas=self.optionExecl(path=path, mode='r', )
        tool = CsvTool()
        tool.optionCsv(path=path.replace('.xlsx','.csv'), encoding=encoding, mode='w',datas=datas)
    def chageCsvToExcelFile(self,path="",encoding='utf-8',sheetName="Sheet1"):
        """
        #将csv文件转成excel文件
        :param path:
        :param encoding:
        :param sheetName:
        :return:
        """
        csvTool=csvFileTool()
        csvArr=csvTool.readCsvData_arrDict(path,encoding=encoding)

        self.optionExecl(path=path.replace('.csv','.xlsx'),sheetName=sheetName,mode='w',datas=csvArr)

    def filter(self,path="",attr="",conditionFunction="",sheetName="Sheet1"):
        filterItems=[]
        items=self.optionExecl(path=path,sheetName=sheetName,mode='r')
        for item in items:
            if conditionFunction(item[attr]):
                filterItems.append(item)
        return filterItems

if __name__ == '__main__':

        # citys=["广东","深圳","珠海","汕头","佛山","韶关市","湛江","肇庆","江门","茂名","惠州","梅州","汕尾","河源","阳江","清远","东莞","中山","潮州","揭阳","云浮",
        #       "福建","福州","厦门","宁德","莆田","泉州","漳州","龙岩","三明","南平"
        #       "江西","南昌","景德","萍乡","九江","新余","鹰潭","赣州","吉安","宜春","抚州","上饶",
        #        "湖南",""
        #       ]
        # for city in citys:
        #     if city.strip() != "":
        #         if city in item:
        #             return True
        # return False
    tool=ExcelTool()
    datas=tool.optionExecl(path='test.xlsx', mode='w', datas=[{'公司1': '1江山'}])
    print(datas)
    # tool.chageCsvToExcelFile('resouse/顺企网_key=高强涤纶.csv', encoding='ANSI')
    # arr=tool.filter(path='resouse/顺企网_key=高强涤纶.xlsx', attr="地址", conditionFunction=test)
    # data = tool.optionExecl(path='resouse/test.xlsx', datas=arr, mode='w')




    # datas=[{'name':'张三','age':10,'资料':"<href='C:/Users/1234567/Desktop/hrefsFile/舒服的耳语.txt'>舒服的耳语.txt"},{'name': '历史', 'age': 12, '资料': 'hrefsFile/舒服的耳语.txt'}]
    #
    # tool.optionExecl(path='resouse/cases.xlsx',sheetName='112',datas=datas,mode='w')
    # data = tool.optionExecl(path='resouse/cases.xlsx',datas=datas,sheetName='112', mode='a')
    # # data=tool.optionExecl(path='resouse/cases.xlsx',datas=datas ,sheetName='112',mode='w')
