import os
import re
import openpyxl
from openpyxl.styles import Font
# from _xhr_tool._utils.CsvTool import CsvTool
# from _xhr_tool._utils.excelTool.csvFileOptionTool import csvFileTool
from utils_xhr.err import ExcelToolException
def _checkExcel(path):
    """检查路径是否是checkExcel"""
    from utils_xhr.io import _checkFile
    _checkFile(path=path,suffix='.xlsx')
def _createNewWorkbook():
    """创建workbook对象"""
    return openpyxl.Workbook()

def _createNewSheet(workbook,sheetName):
    """根据workbook创建sheet对象"""
    workbook.get_sheet_names
    return workbook.create_sheet(sheetName)
def _checkSheetNumber(path,workbook):
    """
    判断workbook对象中，sheet是否有且只有一个，如果不是只有一个，会抛出错误
    """
    if len(workbook.sheetnames) != 1:
        raise ExcelToolException('Excel文件中的Sheet表格有且只能为1个.检测到该文件{}的sheet表格数量不唯一，请删除对于表格。'.format(path))
def _getSheetNameByIndex(workbook,index=0):
    return workbook.sheetnames[index]
def _changeCellToStr(cell):
    """????"""
    if cell.hyperlink:
        cell.value="<href='"+cell.hyperlink.target+"'>"+cell.value
    return cell.value
def _changeStrToCell(cell):
    """????"""
    attrss=re.findall(r'<(.*)>',cell.value)
    if len(attrss)>0:
        arr=[]
        attrs=attrss[0].split(' ')
        for attr in attrs:
            if attr.strip()!='':
                arr.append(attr)
        for ar in arr:
            try:
                tu=ar.split('=')
                if tu[1].startswith("'") and tu[1].endswith("'"):
                    pass
                elif tu[1].startswith("\"") and tu[1].endswith("\""):
                    pass
                else:
                    raise ValueError()
            except:
                raise ValueError("语法错误"+ar+"：无法处理该语句，请确保属性中有等号")
            if tu[0]=='href':############################################################################此处扩展
                cell.hyperlink=tu[1][1:len(tu[1])-1]
                cell.value=re.sub(r'<.*>','',cell.value)
                cell.font=Font(bold=False,italic=False,underline="single",color='0000FF')
            else:
                raise ValueError('语法错误：不存在'+tu[0]+'这个属性,目前只允许以下属性存在：href')


    return cell.value
def _changeStyle(self,cell,fontStyle={}):
    """
    改变文字的样式fontStype
    underline:single | none _添加下划线
    fontColor: 0000FF _蓝色
    size: 11
    bold: True _是否加粗
    name: 宋体 _字体样式
    italic: True   _是否斜体
    strike: False _是否添加删除线
    :param cell:
    :param fontStyle:
    :return:
    """
    #设置默认样式
    if "size" in fontStyle.keys():
        pass
    cell.font = Font(size=fontStyle['size']|100, bold=True, italic=True, underline="single", color='0000FF')
    return cell
def optionExecl(path='',datas=[],mode="r" or "w" or "a",styleRemain=True,isCreateFile=False):
    """操作excel文件。包括读(r) 写(w) 添加(a)
    :param path: 文件路径
    :param datas: 需要写入的数据(w)(a)写与添加模式会用到。数据格式：datas:[{'name':'张三','age':2},{'name':'李四','age':4}]
    :param mode: 模式，目前有3中。r 读模式-用于读取excel文件；w 写模式用于写入文档；a 添加模式用于为文档添加信息。
    :param styleRemain: 是否保留excel的格式。
    """
    if isCreateFile:
        if not path.endswith('.xlsx'):
            raise FileExistsError(path+'文件后缀名错误！不为.xlsx')
    else:
        _checkExcel(path=path)
    if mode=='w':
        wb=_createNewWorkbook()
        sheetName=_getSheetNameByIndex(wb)
        sheet=wb[sheetName]
        #表头的写入
        if len(datas)==0:
            # del wb['Sheet1']  # 删除默认表单
            # wb.save(path)
            wb.close()
            return
        firstArr=list(datas[0].keys())
        for i in range(0,len(firstArr)):
            sheet.cell(row=1,column=i+1,value=firstArr[i])
        #表体的写入
        for j in range(0,len(datas)):
            for i in range(0,len(firstArr)):
                cell=sheet.cell(row=j + 2, column=i + 1, value=datas[j].get(firstArr[i]))
                #对value进行处理
                if type(cell.value)==str:
                    cell=_changeStrToCell(cell)
        wb.save(path)
        wb.close()
    elif mode=='r':
        wb = openpyxl.load_workbook(path)
        _checkSheetNumber(path=path ,workbook=wb)
        sh=wb[wb.sheetnames[0]]
        rows_data = list(sh.rows)
        headLine=[]#头
        datas=[]
        for hc in rows_data.pop(0):
            headLine.append(hc.value)
        for row in rows_data:
            data={}
            for i in range(0,len(row)):
                if styleRemain:
                    data.setdefault(headLine[i], _changeCellToStr(row[i]))
                else:
                    data.setdefault(headLine[i],row[i].value)
            datas.append(data)
        wb.close()
        return datas
    elif mode=='a':
        if not os.path.exists(path):
            optionExecl(path=path, datas=datas, mode="w",isCreateFile=True)
            return
        try:
            oldDatas=optionExecl(path=path, mode="r", styleRemain=True)
        except IndexError:
            optionExecl(path=path, datas=datas, mode="w")
            return
        newDatas = oldDatas + datas
        optionExecl(path=path, datas=newDatas, mode="w")
def getHeader(path=''):
    """
    获得excel文件的表头
    """
    _checkExcel(path)
    wb = openpyxl.load_workbook(path)
    _checkSheetNumber(path=path, workbook=wb)
    sh=wb[wb.sheetnames[0]]
    rows_data = list(sh.rows)
    headLine=[]
    for hc in rows_data.pop(0):
        headLine.append(hc.value)
    return headLine
def changeExeclToCsvFile(path="",encoding='utf-8'):
    """
    #将csv文件转成excel文件。目前只能转换第一个表格
    :param path:需要转换的文件的路径
    :param encoding:读取的excel文件的编码，写入编码默认和读取编码一直
    :return: None
    """
    datas=optionExecl(path=path, mode='r')
    from utils_xhr.io.csvTool import optionCsv
    optionCsv(path=path.replace('.xlsx','.csv'), encoding=encoding, mode='w',datas=datas,isCreateFile=True)
def changeCsvToExcelFile(path="",encoding='utf-8'):
    """
    #将csv文件转成excel文件
    :param path:
    :param encoding:
    :return:
    """
    from utils_xhr.io.csvTool import optionCsv
    datas = optionCsv(path=path, mode='r',encoding=encoding)
    optionExecl(path=path.replace('.csv','.xlsx'),mode='w',datas=datas,isCreateFile=True)

def filter(self,path="",attr="",conditionFunction="",sheetName="Sheet1"):
        filterItems=[]
        items=self.optionExecl(path=path,sheetName=sheetName,mode='r')
        for item in items:
            if conditionFunction(item[attr]):
                filterItems.append(item)
        return filterItems
if __name__ == '__main__':
    # datas=optionExecl(path='test1.xlsx',mode='r',isCreateFile=True)
    # print(datas)
    # changeCsvToExcelFile(path='test.csv')
    changeExeclToCsvFile(path='test.xlsx')